import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference
import os
warnings.filterwarnings('ignore')

# Create reports directory
if not os.path.exists('reports'):
    os.makedirs('reports')
    os.makedirs('reports/charts')

# Excel styling functions
def get_header_style():
    return Font(name='Calibri', size=12, bold=True, color='FFFFFF'), \
           PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid'), \
           Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), \
           Alignment(horizontal='center', vertical='center')

def get_title_style():
    return Font(name='Calibri', size=14, bold=True, color='000000'), \
           PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid'), \
           Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), \
           Alignment(horizontal='center', vertical='center')

def get_data_style():
    return Font(name='Calibri', size=10, color='000000'), \
           PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'), \
           Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), \
           Alignment(horizontal='left', vertical='center')

def get_number_style():
    return Font(name='Calibri', size=10, color='000000'), \
           PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'), \
           Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), \
           Alignment(horizontal='right', vertical='center')

# Load and prepare the data
def load_data(filepath):
    df = pd.read_excel(filepath)
    df_split = df.iloc[:, 0].str.split(',', expand=True)
    df_split.columns = ['Date', 'Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume']
    df_split['Date'] = pd.to_datetime(df_split['Date'])
    for col in ['Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume']:
        df_split[col] = pd.to_numeric(df_split[col], errors='coerce')
    df_split.set_index('Date', inplace=True)
    return df_split

# 1. DESCRIPTIVE ANALYTICS
def descriptive_analytics_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Descriptive Analytics"
    
    # Title
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = "DESCRIPTIVE ANALYTICS REPORT"
    title_font, title_fill, title_border, title_align = get_title_style()
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.border = title_border
    title_cell.alignment = title_align
    
    # Basic Statistics
    row = 3
    headers = ['Metric', 'Value', 'Metric', 'Value']
    header_font, header_fill, header_border, header_align = get_header_style()
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_align
    
    # Data
    data = [
        ['Data Period', f"{df.index.min().strftime('%Y-%m-%d')} to {df.index.max().strftime('%Y-%m-%d')}"],
        ['Total Trading Days', f"{len(df):,}"],
        ['Average Daily Volume', f"{df['Volume'].mean():,.0f}"],
        ['Price Range', f"${df['Low'].min():.2f} - ${df['High'].max():.2f}"],
        ['Average Closing Price', f"${df['Close'].mean():.2f}"]
    ]
    
    data_font, data_fill, data_border, data_align = get_data_style()
    number_font, number_fill, number_border, number_align = get_number_style()
    
    for i, (metric, value) in enumerate(data):
        row += 1
        # Metric
        cell = ws.cell(row=row, column=1, value=metric)
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
        
        # Value
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = number_font
        cell.fill = number_fill
        cell.border = number_border
        cell.alignment = number_align
    
    # Statistics Table
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    subtitle = ws.cell(row=row, column=1, value="Price Statistics")
    subtitle.font = title_font
    subtitle.fill = title_fill
    subtitle.border = title_border
    subtitle.alignment = title_align
    
    row += 1
    stats = df[['Open', 'High', 'Low', 'Close']].describe()
    for r in dataframe_to_rows(stats, index=True, header=True):
        for c, value in enumerate(r, 1):
            if row == 9:  # Header row
                cell = ws.cell(row=row, column=c, value=value)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = header_border
                cell.alignment = header_align
            else:
                cell = ws.cell(row=row, column=c, value=value)
                if isinstance(value, (int, float)):
                    cell.font = number_font
                    cell.fill = number_fill
                    cell.number_format = '#,##0.00'
                else:
                    cell.font = data_font
                    cell.fill = data_fill
                cell.border = data_border
                cell.alignment = data_align
        row += 1
    
    # Adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    wb.save('reports/01_descriptive_analytics.xlsx')
    return "Descriptive Analytics Excel report generated"

# 2. PERFORMANCE ANALYTICS
def performance_analytics_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Performance Analytics"
    
    # Title
    ws.merge_cells('A1:B1')
    title_cell = ws['A1']
    title_cell.value = "PERFORMANCE ANALYTICS REPORT"
    title_font, title_fill, title_border, title_align = get_title_style()
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.border = title_border
    title_cell.alignment = title_align
    
    # Calculate metrics
    df['Daily_Return'] = df['Close'].pct_change()
    df['Cumulative_Return'] = (1 + df['Daily_Return']).cumprod()
    
    total_return = df['Cumulative_Return'].iloc[-1] - 1
    annual_return = (df['Cumulative_Return'].iloc[-1] ** (252/len(df))) - 1
    volatility = df['Daily_Return'].std() * np.sqrt(252)
    
    # Headers
    row = 3
    headers = ['Performance Metric', 'Value']
    header_font, header_fill, header_border, header_align = get_header_style()
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_align
    
    # Data
    data = [
        ['Total Return', total_return],
        ['Annualized Return', annual_return],
        ['Annualized Volatility', volatility]
    ]
    
    data_font, data_fill, data_border, data_align = get_data_style()
    
    for i, (metric, value) in enumerate(data):
        row += 1
        # Metric
        cell = ws.cell(row=row, column=1, value=metric)
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
        
        # Value (as percentage)
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
        cell.number_format = '0.00%'
        
        # Color coding
        if value > 0:
            cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
        else:
            cell.fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    
    wb.save('reports/02_performance_analytics.xlsx')
    return "Performance Analytics Excel report generated"

# 3. TECHNICAL ANALYTICS
def technical_analytics_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Technical Analytics"
    
    # Title
    ws.merge_cells('A1:B1')
    title_cell = ws['A1']
    title_cell.value = "TECHNICAL ANALYTICS REPORT"
    title_font, title_fill, title_border, title_align = get_title_style()
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.border = title_border
    title_cell.alignment = title_align
    
    # Calculate indicators
    df['SMA_50'] = df['Close'].rolling(window=50).mean()
    
    def calculate_rsi(data, window=14):
        delta = data.diff()
        gain = (delta.where(delta > 0, 0)).rolling(window=window).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(window=window).mean()
        rs = gain / loss
        rsi = 100 - (100 / (1 + rs))
        return rsi
    
    df['RSI'] = calculate_rsi(df['Close'])
    
    current_price = df['Close'].iloc[-1]
    current_sma = df['SMA_50'].iloc[-1]
    current_rsi = df['RSI'].iloc[-1]
    
    # Headers
    row = 3
    headers = ['Technical Indicator', 'Current Value']
    header_font, header_fill, header_border, header_align = get_header_style()
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_align
    
    # Data
    data = [
        ['Current Price', current_price],
        ['50-day SMA', current_sma],
        ['RSI (14-day)', current_rsi]
    ]
    
    data_font, data_fill, data_border, data_align = get_data_style()
    
    for i, (metric, value) in enumerate(data):
        row += 1
        # Metric
        cell = ws.cell(row=row, column=1, value=metric)
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
        
        # Value
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
        
        if 'Price' in metric or 'SMA' in metric:
            cell.number_format = '$#,##0.00'
        elif 'RSI' in metric:
            cell.number_format = '0.0'
    
    # Signal
    row += 2
    signal = "BULLISH" if current_price > current_sma else "BEARISH"
    signal_cell = ws.cell(row=row, column=1, value="Trading Signal")
    signal_cell.font = data_font
    signal_cell.fill = data_fill
    signal_cell.border = data_border
    signal_cell.alignment = data_align
    
    signal_value = ws.cell(row=row, column=2, value=signal)
    signal_value.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    if signal == "BULLISH":
        signal_value.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    else:
        signal_value.fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
    signal_value.border = data_border
    signal_value.alignment = header_align
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    
    wb.save('reports/03_technical_analytics.xlsx')
    return "Technical Analytics Excel report generated"

# 4. RISK ANALYTICS
def risk_analytics_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Analytics"
    
    # Title
    ws.merge_cells('A1:B1')
    title_cell = ws['A1']
    title_cell.value = "RISK ANALYTICS REPORT"
    title_font, title_fill, title_border, title_align = get_title_style()
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.border = title_border
    title_cell.alignment = title_align
    
    # Calculate risk metrics
    returns = df['Close'].pct_change().dropna()
    var_95 = np.percentile(returns, 5)
    
    cumulative = (1 + returns).cumprod()
    running_max = cumulative.expanding().max()
    drawdown = (cumulative - running_max) / running_max
    max_drawdown = drawdown.min()
    
    risk_free_rate = 0.02
    excess_returns = returns - risk_free_rate/252
    sharpe_ratio = np.sqrt(252) * excess_returns.mean() / returns.std()
    
    # Headers
    row = 3
    headers = ['Risk Metric', 'Value']
    header_font, header_fill, header_border, header_align = get_header_style()
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_align
    
    # Data
    data = [
        ['Value at Risk (95% daily)', var_95],
        ['Maximum Drawdown', max_drawdown],
        ['Sharpe Ratio', sharpe_ratio]
    ]
    
    data_font, data_fill, data_border, data_align = get_data_style()
    
    for i, (metric, value) in enumerate(data):
        row += 1
        # Metric
        cell = ws.cell(row=row, column=1, value=metric)
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
        
        # Value
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
        
        if 'Ratio' in metric:
            cell.number_format = '0.00'
        else:
            cell.number_format = '0.00%'
        
        # Color coding for risk metrics
        if 'Drawdown' in metric or 'VaR' in metric:
            cell.fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
        elif 'Sharpe' in metric:
            if value > 1:
                cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
            elif value > 0:
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    
    wb.save('reports/04_risk_analytics.xlsx')
    return "Risk Analytics Excel report generated"

# 5. TIME SERIES ANALYTICS
def time_series_analytics_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Series Analytics"
    
    # Title
    ws.merge_cells('A1:B1')
    title_cell = ws['A1']
    title_cell.value = "TIME SERIES ANALYTICS REPORT"
    title_font, title_fill, title_border, title_align = get_title_style()
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.border = title_border
    title_cell.alignment = title_align
    
    # Calculate monthly patterns
    monthly_returns = df['Close'].resample('M').last().pct_change()
    avg_monthly_return = monthly_returns.groupby(monthly_returns.index.month).mean()
    
    best_month = avg_monthly_return.idxmax()
    worst_month = avg_monthly_return.idxmin()
    
    yearly_avg = df['Close'].resample('Y').mean()
    recent_trend = yearly_avg.tail(5).pct_change().mean()
    
    # Headers
    row = 3
    headers = ['Time Series Metric', 'Value']
    header_font, header_fill, header_border, header_align = get_header_style()
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_align
    
    # Data
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    data = [
        ['Best Performing Month', f"{months[best_month-1]} ({avg_monthly_return.max():.2%})"],
        ['Worst Performing Month', f"{months[worst_month-1]} ({avg_monthly_return.min():.2%})"],
        ['Recent 5-Year Trend', f"{recent_trend:.2%} per year"]
    ]
    
    data_font, data_fill, data_border, data_align = get_data_style()
    
    for i, (metric, value) in enumerate(data):
        row += 1
        # Metric
        cell = ws.cell(row=row, column=1, value=metric)
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
        
        # Value
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    wb.save('reports/05_time_series_analytics.xlsx')
    return "Time Series Analytics Excel report generated"

# 6. VOLATILITY ANALYTICS
def volatility_analytics_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Volatility Analytics"
    
    # Title
    ws.merge_cells('A1:B1')
    title_cell = ws['A1']
    title_cell.value = "VOLATILITY ANALYTICS REPORT"
    title_font, title_fill, title_border, title_align = get_title_style()
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.border = title_border
    title_cell.alignment = title_align
    
    # Calculate volatility metrics
    df['Daily_Return'] = df['Close'].pct_change()
    df['Volatility_30d'] = df['Daily_Return'].rolling(window=30).std() * np.sqrt(252)
    
    current_vol = df['Volatility_30d'].iloc[-1]
    avg_vol = df['Volatility_30d'].mean()
    high_vol_threshold = df['Volatility_30d'].quantile(0.8)
    
    # Headers
    row = 3
    headers = ['Volatility Metric', 'Value']
    header_font, header_fill, header_border, header_align = get_header_style()
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_align
    
    # Data
    data = [
        ['Current 30-day Volatility', current_vol],
        ['Average Volatility', avg_vol],
        ['High Volatility Threshold (80th %)', high_vol_threshold]
    ]
    
    data_font, data_fill, data_border, data_align = get_data_style()
    
    for i, (metric, value) in enumerate(data):
        row += 1
        # Metric
        cell = ws.cell(row=row, column=1, value=metric)
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
        
        # Value
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
        cell.number_format = '0.00%'
        
        # Color coding for volatility
        if 'Current' in metric:
            if value > high_vol_threshold:
                cell.fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
    
    # Market condition
    row += 2
    condition = "HIGH VOLATILITY" if current_vol > high_vol_threshold else "NORMAL VOLATILITY"
    condition_cell = ws.cell(row=row, column=1, value="Market Condition")
    condition_cell.font = data_font
    condition_cell.fill = data_fill
    condition_cell.border = data_border
    condition_cell.alignment = data_align
    
    condition_value = ws.cell(row=row, column=2, value=condition)
    condition_value.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    if condition == "HIGH VOLATILITY":
        condition_value.fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
    else:
        condition_value.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    condition_value.border = data_border
    condition_value.alignment = header_align
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    
    wb.save('reports/06_volatility_analytics.xlsx')
    return "Volatility Analytics Excel report generated"

# 7. PREDICTIVE ANALYTICS
def predictive_analytics_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Predictive Analytics"
    
    # Title
    ws.merge_cells('A1:B1')
    title_cell = ws['A1']
    title_cell.value = "PREDICTIVE ANALYTICS REPORT"
    title_font, title_fill, title_border, title_align = get_title_style()
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.border = title_border
    title_cell.alignment = title_align
    
    # Calculate moving averages
    df['SMA_20'] = df['Close'].rolling(window=20).mean()
    df['SMA_50'] = df['Close'].rolling(window=50).mean()
    
    current_price = df['Close'].iloc[-1]
    sma_20 = df['SMA_20'].iloc[-1]
    sma_50 = df['SMA_50'].iloc[-1]
    
    # Determine prediction
    if sma_20 > sma_50 and df['SMA_20'].iloc[-2] <= df['SMA_50'].iloc[-2]:
        prediction = "GOLDEN CROSS - Bullish signal"
        prediction_color = '70AD47'
    elif sma_20 < sma_50 and df['SMA_20'].iloc[-2] >= df['SMA_50'].iloc[-2]:
        prediction = "DEATH CROSS - Bearish signal"
        prediction_color = 'E74C3C'
    elif sma_20 > sma_50:
        prediction = "Bullish trend continuation"
        prediction_color = 'C6E0B4'
    else:
        prediction = "Bearish trend continuation"
        prediction_color = 'F8CBAD'
    
    # Headers
    row = 3
    headers = ['Predictive Indicator', 'Current Value']
    header_font, header_fill, header_border, header_align = get_header_style()
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_align
    
    # Data
    data = [
        ['20-day SMA', sma_20],
        ['50-day SMA', sma_50]
    ]
    
    data_font, data_fill, data_border, data_align = get_data_style()
    
    for i, (metric, value) in enumerate(data):
        row += 1
        # Metric
        cell = ws.cell(row=row, column=1, value=metric)
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
        
        # Value
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
        cell.number_format = '$#,##0.00'
    
    # Prediction
    row += 2
    pred_cell = ws.cell(row=row, column=1, value="Prediction")
    pred_cell.font = data_font
    pred_cell.fill = data_fill
    pred_cell.border = data_border
    pred_cell.alignment = data_align
    
    pred_value = ws.cell(row=row, column=2, value=prediction)
    pred_value.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    pred_value.fill = PatternFill(start_color=prediction_color, end_color=prediction_color, fill_type='solid')
    pred_value.border = data_border
    pred_value.alignment = header_align
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    
    wb.save('reports/07_predictive_analytics.xlsx')
    return "Predictive Analytics Excel report generated"

# 8. TRADING STRATEGY ANALYTICS
def trading_strategy_analytics_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Trading Strategy Analytics"
    
    # Title
    ws.merge_cells('A1:B1')
    title_cell = ws['A1']
    title_cell.value = "TRADING STRATEGY ANALYTICS REPORT"
    title_font, title_fill, title_border, title_align = get_title_style()
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.border = title_border
    title_cell.alignment = title_align
    
    # Calculate strategy performance
    df['Daily_Return'] = df['Close'].pct_change()
    df['SMA_20'] = df['Close'].rolling(window=20).mean()
    
    df['Signal'] = np.where(df['Close'] > df['SMA_20'], 1, -1)
    df['Strategy_Return'] = df['Signal'].shift(1) * df['Daily_Return']
    
    strategy_cumulative = (1 + df['Strategy_Return'].dropna()).cumprod()
    buy_hold_cumulative = (1 + df['Daily_Return'].dropna()).cumprod()
    
    strategy_return = strategy_cumulative.iloc[-1] - 1
    buy_hold_return = buy_hold_cumulative.iloc[-1] - 1
    outperformance = strategy_return - buy_hold_return
    
    # Headers
    row = 3
    headers = ['Strategy Performance', 'Return']
    header_font, header_fill, header_border, header_align = get_header_style()
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_align
    
    # Data
    data = [
        ['SMA Strategy Total Return', strategy_return],
        ['Buy & Hold Return', buy_hold_return],
        ['Strategy Outperformance', outperformance]
    ]
    
    data_font, data_fill, data_border, data_align = get_data_style()
    
    for i, (metric, value) in enumerate(data):
        row += 1
        # Metric
        cell = ws.cell(row=row, column=1, value=metric)
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
        
        # Value
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
        cell.number_format = '0.00%'
        
        # Color coding
        if 'Outperformance' in metric:
            if value > 0:
                cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
    
    # Result
    row += 2
    result = "Strategy BEATS Buy & Hold" if strategy_return > buy_hold_return else "Buy & Hold BEATS Strategy"
    result_color = '70AD47' if strategy_return > buy_hold_return else 'F8CBAD'
    
    result_cell = ws.cell(row=row, column=1, value="Final Result")
    result_cell.font = data_font
    result_cell.fill = data_fill
    result_cell.border = data_border
    result_cell.alignment = data_align
    
    result_value = ws.cell(row=row, column=2, value=result)
    result_value.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    result_value.fill = PatternFill(start_color=result_color, end_color=result_color, fill_type='solid')
    result_value.border = data_border
    result_value.alignment = header_align
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    wb.save('reports/08_trading_strategy_analytics.xlsx')
    return "Trading Strategy Analytics Excel report generated"

def main():
    print("GENERATING EXCEL BUSINESS ANALYTICS REPORTS")
    print("=" * 50)
    print()
    
    # Load data
    df = load_data('DevicesData.xlsx')
    
    # Generate Excel reports
    reports = []
    reports.append(descriptive_analytics_excel(df))
    reports.append(performance_analytics_excel(df))
    reports.append(technical_analytics_excel(df))
    reports.append(risk_analytics_excel(df))
    reports.append(time_series_analytics_excel(df))
    reports.append(volatility_analytics_excel(df))
    reports.append(predictive_analytics_excel(df))
    reports.append(trading_strategy_analytics_excel(df))
    
    print("✓ All 8 Excel reports generated successfully!")
    print("Generated Files:")
    print("  - reports/01_descriptive_analytics.xlsx")
    print("  - reports/02_performance_analytics.xlsx")
    print("  - reports/03_technical_analytics.xlsx")
    print("  - reports/04_risk_analytics.xlsx")
    print("  - reports/05_time_series_analytics.xlsx")
    print("  - reports/06_volatility_analytics.xlsx")
    print("  - reports/07_predictive_analytics.xlsx")
    print("  - reports/08_trading_strategy_analytics.xlsx")
    print()
    print("Features:")
    print("  - Professional color schemes")
    print("  - Conditional formatting")
    print("  - Proper number formatting")
    print("  - Bordered tables")
    print("  - Centered headers")
    print("  - Color-coded results")

if __name__ == "__main__":
    main()
